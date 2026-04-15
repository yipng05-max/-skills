#!/usr/bin/env python3
"""
根据方法编码结果生成 Excel 方法矩阵 + Markdown 趋势报告。

输入：
    --papers /path/to/papers.json          # fetch_journal_papers.py 产出
    --classifications /path/to/classifications.json  # LLM 编码结果
    --out-dir /path/to/output/

classifications.json 格式（见 README 中 schema）：
{
  "classifications": [
    {
      "openalex_id": "https://openalex.org/Wxxxx",
      "l1_primary": "qual.ethnography",
      "l1_secondary": "ai.llm",
      "data_type": "...",
      ...（完整字段见 excel_schema.md）
      "teaching_candidate": true,
      "l3_operational_summary": "..."
    },
    ...
  ]
}
"""

from __future__ import annotations

import argparse
import json
import re
from collections import Counter, defaultdict
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

MAIN_COLS = [
    "paper_id", "journal", "discipline", "tier", "publication_date",
    "title", "authors", "doi", "cited_by_count",
    "l1_primary", "l1_secondary",
    "data_type", "data_source", "sample_size", "time_window_data",
    "geographic_scope", "analytic_tools", "validity_strategy",
    "ai_involved", "ai_category", "ai_subcategory",
    "methodological_innovation", "teaching_candidate",
    "l3_operational_summary", "notes",
]

AI_COLS = [
    "paper_id", "journal", "discipline", "title", "doi",
    "ai_category", "ai_subcategory", "ai_role",
    "ai_model_used", "validation_of_ai", "ethics_statement", "teaching_hook",
]

DISC_COLS = [
    "discipline", "n_papers", "top_3_l1_methods",
    "n_ai_papers", "ai_percentage",
    "n_methodological_innovations", "n_teaching_candidates",
    "disciplinary_notes",
]

JOURNAL_COLS = [
    "journal", "abbr", "issn", "discipline", "tier",
    "n_papers_in_window", "date_range_of_papers", "skipped_reason",
]


def style_header(ws, n_cols: int) -> None:
    fill = PatternFill("solid", fgColor="305496")
    font = Font(bold=True, color="FFFFFF")
    for col in range(1, n_cols + 1):
        c = ws.cell(row=1, column=col)
        c.fill = fill
        c.font = font
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.freeze_panes = "A2"


def autosize(ws, max_width: int = 60) -> None:
    for col_cells in ws.columns:
        length = max((len(str(c.value or "")) for c in col_cells), default=10)
        ws.column_dimensions[col_cells[0].column_letter].width = min(
            max(length + 2, 12), max_width
        )


def build_excel(papers_by_id, classifications_by_id, journals_meta, out_path):
    wb = Workbook()

    # Sheet 1: methods_matrix
    ws1 = wb.active
    ws1.title = "methods_matrix"
    ws1.append(MAIN_COLS)
    style_header(ws1, len(MAIN_COLS))
    for i, (oid, p) in enumerate(papers_by_id.items(), start=1):
        c = classifications_by_id.get(oid, {})
        row = [
            f"{i:03d}",
            p.get("journal", ""),
            p.get("discipline", ""),
            c.get("tier", journals_meta.get(p.get("journal", ""), {}).get("tier", "")),
            p.get("publication_date", ""),
            p.get("title", ""),
            p.get("authors", ""),
            p.get("doi", ""),
            p.get("cited_by_count", 0),
            c.get("l1_primary", "?"),
            c.get("l1_secondary", ""),
            c.get("data_type", "?"),
            c.get("data_source", "?"),
            c.get("sample_size", "?"),
            c.get("time_window_data", "?"),
            c.get("geographic_scope", "?"),
            c.get("analytic_tools", "?"),
            c.get("validity_strategy", "?"),
            c.get("ai_involved", "no"),
            c.get("ai_category", "n/a"),
            c.get("ai_subcategory", "n/a"),
            c.get("methodological_innovation", "no"),
            c.get("teaching_candidate", "no"),
            c.get("l3_operational_summary", ""),
            c.get("notes", ""),
        ]
        ws1.append(row)
    autosize(ws1)

    # Sheet 2: ai_papers
    ws2 = wb.create_sheet("ai_papers")
    ws2.append(AI_COLS)
    style_header(ws2, len(AI_COLS))
    ai_rows = []
    for i, (oid, p) in enumerate(papers_by_id.items(), start=1):
        c = classifications_by_id.get(oid, {})
        if str(c.get("ai_involved", "no")).lower() != "yes":
            continue
        ai_rows.append([
            f"{i:03d}",
            p.get("journal", ""),
            p.get("discipline", ""),
            p.get("title", ""),
            p.get("doi", ""),
            c.get("ai_category", ""),
            c.get("ai_subcategory", ""),
            c.get("ai_role", ""),
            c.get("ai_model_used", ""),
            c.get("validation_of_ai", ""),
            c.get("ethics_statement", ""),
            c.get("teaching_hook", ""),
        ])
    for r in ai_rows:
        ws2.append(r)
    autosize(ws2)

    # Sheet 3: by_discipline_summary
    ws3 = wb.create_sheet("by_discipline_summary")
    ws3.append(DISC_COLS)
    style_header(ws3, len(DISC_COLS))
    per_disc = defaultdict(list)
    for oid, p in papers_by_id.items():
        per_disc[p.get("discipline", "?")].append(
            (p, classifications_by_id.get(oid, {}))
        )
    for disc, pairs in sorted(per_disc.items()):
        n = len(pairs)
        l1_counter = Counter()
        for _, c in pairs:
            if c.get("l1_primary"):
                l1_counter[c["l1_primary"]] += 1
            if c.get("l1_secondary"):
                l1_counter[c["l1_secondary"]] += 1
        top3 = "; ".join(f"{k}({v})" for k, v in l1_counter.most_common(3))
        n_ai = sum(1 for _, c in pairs if str(c.get("ai_involved", "no")).lower() == "yes")
        n_innov = sum(1 for _, c in pairs if str(c.get("methodological_innovation", "no")).lower() == "yes")
        n_teach = sum(1 for _, c in pairs if str(c.get("teaching_candidate", "no")).lower() == "yes")
        ws3.append([
            disc, n, top3, n_ai,
            f"{(n_ai / n * 100) if n else 0:.1f}%",
            n_innov, n_teach, "",
        ])
    autosize(ws3)

    # Sheet 4: journals_meta
    ws4 = wb.create_sheet("journals_meta")
    ws4.append(JOURNAL_COLS)
    style_header(ws4, len(JOURNAL_COLS))
    per_j = defaultdict(list)
    for oid, p in papers_by_id.items():
        per_j[p.get("journal", "?")].append(p.get("publication_date", ""))
    for jname, meta in journals_meta.items():
        dates = [d for d in per_j.get(jname, []) if d]
        date_range = ""
        if dates:
            date_range = f"{min(dates)} → {max(dates)}"
        ws4.append([
            jname,
            meta.get("abbr", ""),
            "; ".join(meta.get("issn", [])),
            meta.get("discipline", ""),
            meta.get("tier", ""),
            len(per_j.get(jname, [])),
            date_range,
            "" if per_j.get(jname) else "no papers returned in window",
        ])
    autosize(ws4)

    wb.save(out_path)


def build_markdown(papers_by_id, classifications_by_id, meta, journals_meta, out_path):
    """生成简版 Markdown 报告骨架。深度内容由主 agent 根据数据补写。"""
    total = len(papers_by_id)
    ai_papers = [oid for oid, c in classifications_by_id.items()
                 if str(c.get("ai_involved", "no")).lower() == "yes"]
    ai_tool = [oid for oid in ai_papers
               if classifications_by_id[oid].get("ai_category") == "tool"]
    ai_object = [oid for oid in ai_papers
                 if classifications_by_id[oid].get("ai_category") == "object"]
    ai_hybrid = [oid for oid in ai_papers
                 if classifications_by_id[oid].get("ai_category") == "hybrid"]

    # L1 频次
    l1_counter = Counter()
    for c in classifications_by_id.values():
        if c.get("l1_primary"):
            l1_counter[c["l1_primary"]] += 1
        if c.get("l1_secondary"):
            l1_counter[c["l1_secondary"]] += 1

    # 按学科分组
    per_disc = defaultdict(list)
    for oid, p in papers_by_id.items():
        per_disc[p.get("discipline", "?")].append(
            (p, classifications_by_id.get(oid, {}))
        )

    lines = []
    lines.append("# 全球人文社科顶刊研究方法扫描报告")
    lines.append("")
    lines.append(f"> 时间窗口：{meta.get('since', '?')} → {datetime.today().date().isoformat()}"
                 f"（近 {meta.get('months', '?')} 个月）")
    lines.append(f"> 覆盖期刊：{len(journals_meta)} 本，涉及 {len(per_disc)} 个学科")
    lines.append(f"> 共纳入论文：{total} 篇（去重后）")
    lines.append(f"> 生成时间：{datetime.now().strftime('%Y-%m-%d %H:%M')}")
    lines.append("")
    lines.append("## 摘要")
    lines.append("")
    lines.append("_[此处由主 agent 根据后续数据补写 300-500 字执行摘要，"
                 "包含：(1) 方法分布整体画面；(2) 三个最值得关注的新动向；"
                 "(3) AI 方法总量与形态；(4) 给课程策划人员的 2-3 条建议。]_")
    lines.append("")

    lines.append("## 一、方法家族分布全景")
    lines.append("")
    lines.append("### 1.1 L1 方法标签频次（全库）")
    lines.append("")
    lines.append("| 方法标签 | 频次 | 占比 |")
    lines.append("|---|---:|---:|")
    for tag, n in l1_counter.most_common(25):
        lines.append(f"| `{tag}` | {n} | {n/total*100:.1f}% |")
    lines.append("")

    lines.append("### 1.2 学科 × L1 交叉（频次 ≥ 3）")
    lines.append("")
    disc_tag_matrix = defaultdict(Counter)
    for disc, pairs in per_disc.items():
        for _, c in pairs:
            for k in ("l1_primary", "l1_secondary"):
                if c.get(k):
                    disc_tag_matrix[disc][c[k]] += 1
    lines.append("| 学科 | 主要方法（频次） |")
    lines.append("|---|---|")
    for disc in sorted(disc_tag_matrix):
        cells = "; ".join(f"`{t}`×{n}" for t, n in disc_tag_matrix[disc].most_common()
                          if n >= 3)
        if cells:
            lines.append(f"| {disc} | {cells} |")
    lines.append("")

    lines.append("### 1.3 各学科方法画像")
    lines.append("")
    for disc, pairs in sorted(per_disc.items()):
        lines.append(f"#### {disc}（{len(pairs)} 篇）")
        lines.append("")
        lines.append("_[主 agent 补 150-200 字：主导方法、独特性、转向迹象]_")
        lines.append("")

    lines.append("## 二、新兴方法（Emerging Methods）")
    lines.append("")
    lines.append("_[主 agent 根据 `methodological_innovation=yes` 的论文以及 "
                 "`emerging.*` 前缀的标签，识别 3-6 个新兴方法并按 `report_template.md` "
                 "第二章的格式展开。]_")
    lines.append("")

    lines.append("## 三、AI 方法专章")
    lines.append("")
    lines.append(f"- AI 相关论文总数：**{len(ai_papers)}**（占全库 "
                 f"{len(ai_papers)/total*100:.1f}%）")
    lines.append(f"- AI-as-tool：{len(ai_tool)}")
    lines.append(f"- AI-as-object：{len(ai_object)}")
    lines.append(f"- Hybrid：{len(ai_hybrid)}")
    lines.append("")

    ai_sub_counter = Counter()
    for oid in ai_papers:
        sub = classifications_by_id[oid].get("ai_subcategory")
        if sub:
            ai_sub_counter[sub] += 1
    if ai_sub_counter:
        lines.append("### AI 子类分布")
        lines.append("")
        lines.append("| 子类 | 频次 |")
        lines.append("|---|---:|")
        for sub, n in ai_sub_counter.most_common():
            lines.append(f"| `{sub}` | {n} |")
        lines.append("")

    lines.append("### 3.1 AI-as-tool")
    lines.append("")
    lines.append("_[主 agent 按 5 子类展开，每类含：频次/学科分布、2-3 典型论文一句话、方法论争论点。]_")
    lines.append("")
    lines.append("### 3.2 AI-as-object")
    lines.append("")
    lines.append("_[主 agent 按 5 子类展开。]_")
    lines.append("")
    lines.append("### 3.3 Hybrid 与前沿张力")
    lines.append("")
    lines.append("_[主 agent 逐一讨论 hybrid 论文。]_")
    lines.append("")
    lines.append("### 3.4 AI 方法的伦理与效度挑战")
    lines.append("")
    lines.append("_[主 agent 讨论同意/隐私、黑箱审稿回应、人-AI 编码一致性。]_")
    lines.append("")

    lines.append("## 四、方法论文专栏")
    lines.append("")
    methods_papers = [oid for oid, p in papers_by_id.items()
                      if journals_meta.get(p.get("journal", ""), {}).get("tier") == "methods"]
    if methods_papers:
        lines.append(f"方法期刊本期共 {len(methods_papers)} 篇。值得关注：")
        lines.append("")
        for oid in methods_papers[:10]:
            p = papers_by_id[oid]
            lines.append(f"- **{p.get('title', '')}** ({p.get('journal', '')}, "
                         f"{p.get('publication_date', '')}) — {p.get('doi', '')}")
        lines.append("")
    else:
        lines.append("本期方法期刊样本不足。")
        lines.append("")

    lines.append("## 五、教学转化建议")
    lines.append("")
    lines.append("### 5.1 可立即上线的新模块")
    lines.append("")
    lines.append("_[主 agent 根据 teaching_candidate=yes 的论文拟 3-5 个教学模块，格式见 report_template.md 5.1]_")
    lines.append("")
    lines.append("### 5.2 已有课程需要更新的点")
    lines.append("")
    lines.append("_[主 agent 补写]_")
    lines.append("")
    lines.append("### 5.3 跨院系共建建议")
    lines.append("")
    lines.append("_[主 agent 补写]_")
    lines.append("")

    lines.append("## 六、值得关注的方法缺口")
    lines.append("")
    lines.append("_[主 agent 给出前瞻性判断：本期未见但可能在未来 12-24 个月进入主流的方法。]_")
    lines.append("")

    lines.append("## 七、数据与局限")
    lines.append("")
    lines.append(f"- 数据源：OpenAlex，时间窗 {meta.get('since', '?')} → "
                 f"{datetime.today().date().isoformat()}")
    lines.append("- 方法提取基于摘要+关键词+概念元数据，少数论文需读全文才能准确归类")
    lines.append("- 单篇论文标签仅供扫描用，不能替代精读")
    lines.append("")

    lines.append("## 附录 A · 期刊覆盖")
    lines.append("")
    lines.append("| 期刊 | 学科 | 层级 | 本期论文数 |")
    lines.append("|---|---|---|---:|")
    per_j = Counter()
    for p in papers_by_id.values():
        per_j[p.get("journal", "?")] += 1
    for jname, meta_j in journals_meta.items():
        lines.append(f"| {jname} | {meta_j.get('discipline', '')} "
                     f"| {meta_j.get('tier', '')} | {per_j.get(jname, 0)} |")
    lines.append("")

    lines.append("## 附录 B · AI 相关论文完整清单")
    lines.append("")
    for oid in ai_papers:
        p = papers_by_id[oid]
        c = classifications_by_id[oid]
        lines.append(f"- [{c.get('ai_category', '')}] [{p.get('journal', '')}] "
                     f"{p.get('title', '')} — {p.get('authors', '')} — "
                     f"{p.get('doi', '')}")
    lines.append("")

    out_path.write_text("\n".join(lines), encoding="utf-8")


def main() -> int:
    ap = argparse.ArgumentParser()
    ap.add_argument("--papers", required=True)
    ap.add_argument("--classifications", required=True)
    ap.add_argument("--journals-file", required=True)
    ap.add_argument("--out-dir", required=True)
    args = ap.parse_args()

    papers_data = json.loads(Path(args.papers).read_text(encoding="utf-8"))
    class_data = json.loads(Path(args.classifications).read_text(encoding="utf-8"))
    journals_cfg = json.loads(Path(args.journals_file).read_text(encoding="utf-8"))

    papers_by_id = {p["openalex_id"]: p for p in papers_data["papers"]}
    classifications_by_id = {
        c["openalex_id"]: c for c in class_data.get("classifications", [])
    }

    journals_meta = {}
    for disc, items in journals_cfg["disciplines"].items():
        for j in items:
            journals_meta[j["name"]] = {**j, "discipline": disc}

    out_dir = Path(args.out_dir)
    out_dir.mkdir(parents=True, exist_ok=True)
    ts = datetime.now().strftime("%Y%m%d_%H%M")
    xlsx_path = out_dir / f"methods_matrix_{ts}.xlsx"
    md_path = out_dir / f"methods_report_{ts}.md"

    build_excel(papers_by_id, classifications_by_id, journals_meta, xlsx_path)
    build_markdown(papers_by_id, classifications_by_id,
                   papers_data.get("meta", {}), journals_meta, md_path)

    print(f"✓ Excel 已保存：{xlsx_path}")
    print(f"✓ Markdown 已保存：{md_path}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
