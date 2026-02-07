#!/usr/bin/env python3
"""Export paper analysis results to Excel files.

Usage:
  Single paper:
    python3 export_excel.py --mode single --output /path/to/output.xlsx --json '<json_data>'

  Multiple papers with comparison:
    python3 export_excel.py --mode multi --output /path/to/output.xlsx --json '<json_data>'

JSON format for single paper:
{
  "title": "论文标题",
  "author": "作者",
  "source": "期刊/来源",
  "year": "年份",
  "elements": {
    "研究背景": "...",
    "研究问题": "...",
    "研究结论": "...",
    "文献综合": "...",
    "文献批评": "...",
    "研究方法": "...",
    "理论视角与理论框架": "...",
    "一致性发现": "...",
    "不一致性发现": "...",
    "研究贡献": "...",
    "研究不足": "...",
    "未来研究展望": "..."
  }
}

JSON format for multiple papers:
{
  "papers": [ ...array of single paper objects... ],
  "summary": {
    "paper_labels": ["论文A简称", "论文B简称", ...],
    "elements": {
      "研究背景": ["论文A摘要", "论文B摘要", ...],
      ...
    }
  }
}
"""

import argparse
import json
import sys
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

ELEMENT_NAMES = [
    "1. 研究背景",
    "2. 研究问题",
    "3. 研究结论",
    "4. 文献综合",
    "5. 文献批评",
    "6. 研究方法",
    "7. 理论视角与理论框架",
    "8. 一致性发现",
    "9. 不一致性发现",
    "10. 研究贡献",
    "11. 研究不足",
    "12. 未来研究展望",
]

ELEMENT_KEYS = [
    "研究背景", "研究问题", "研究结论", "文献综合", "文献批评",
    "研究方法", "理论视角与理论框架", "一致性发现", "不一致性发现",
    "研究贡献", "研究不足", "未来研究展望",
]

# Styles
HEADER_FONT = Font(name="Microsoft YaHei", bold=True, size=11, color="FFFFFF")
HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
ELEMENT_FONT = Font(name="Microsoft YaHei", bold=True, size=10)
ELEMENT_FILL = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
CONTENT_FONT = Font(name="Microsoft YaHei", size=10)
META_FONT = Font(name="Microsoft YaHei", bold=True, size=12)
THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)
WRAP_ALIGNMENT = Alignment(wrap_text=True, vertical="top")
HEADER_ALIGNMENT = Alignment(wrap_text=True, vertical="center", horizontal="center")


def style_cell(cell, font=None, fill=None, alignment=None, border=None):
    if font:
        cell.font = font
    if fill:
        cell.fill = fill
    if alignment:
        cell.alignment = alignment
    if border:
        cell.border = border


def write_single_sheet(ws, paper):
    """Write a single paper analysis to a worksheet."""
    ws.column_dimensions["A"].width = 25
    ws.column_dimensions["B"].width = 90

    # Title row
    ws.merge_cells("A1:B1")
    title_cell = ws["A1"]
    title_cell.value = f"论文拆解：《{paper['title']}》"
    style_cell(title_cell, font=Font(name="Microsoft YaHei", bold=True, size=14),
               alignment=Alignment(horizontal="center", vertical="center"))
    ws.row_dimensions[1].height = 30

    # Meta row
    ws.merge_cells("A2:B2")
    meta_cell = ws["A2"]
    meta_cell.value = f"作者: {paper.get('author', '')}  |  期刊/来源: {paper.get('source', '')}  |  年份: {paper.get('year', '')}"
    style_cell(meta_cell, font=Font(name="Microsoft YaHei", size=10, italic=True),
               alignment=Alignment(horizontal="center", vertical="center"))
    ws.row_dimensions[2].height = 22

    # Header row
    row = 4
    for col, header in enumerate(["阅读要素", "内容"], 1):
        cell = ws.cell(row=row, column=col, value=header)
        style_cell(cell, font=HEADER_FONT, fill=HEADER_FILL,
                   alignment=HEADER_ALIGNMENT, border=THIN_BORDER)
    ws.row_dimensions[row].height = 22

    # Data rows
    elements = paper.get("elements", {})
    for i, (name, key) in enumerate(zip(ELEMENT_NAMES, ELEMENT_KEYS)):
        row = 5 + i
        cell_a = ws.cell(row=row, column=1, value=name)
        style_cell(cell_a, font=ELEMENT_FONT, fill=ELEMENT_FILL,
                   alignment=Alignment(wrap_text=True, vertical="top"),
                   border=THIN_BORDER)

        cell_b = ws.cell(row=row, column=2, value=elements.get(key, ""))
        style_cell(cell_b, font=CONTENT_FONT, alignment=WRAP_ALIGNMENT,
                   border=THIN_BORDER)

        # Auto row height based on content length
        content_len = len(elements.get(key, ""))
        ws.row_dimensions[row].height = max(40, min(150, content_len * 0.6))


def write_summary_sheet(ws, summary):
    """Write the multi-paper comparison summary to a worksheet."""
    labels = summary.get("paper_labels", [])
    n_papers = len(labels)

    ws.column_dimensions["A"].width = 25
    for i in range(n_papers):
        col_letter = chr(ord("B") + i)
        ws.column_dimensions[col_letter].width = 40

    # Title
    end_col = chr(ord("A") + n_papers)
    ws.merge_cells(f"A1:{end_col}1")
    title_cell = ws["A1"]
    title_cell.value = "多篇论文横向对比汇总"
    style_cell(title_cell, font=Font(name="Microsoft YaHei", bold=True, size=14),
               alignment=Alignment(horizontal="center", vertical="center"))
    ws.row_dimensions[1].height = 30

    # Header row
    row = 3
    cell = ws.cell(row=row, column=1, value="阅读要素")
    style_cell(cell, font=HEADER_FONT, fill=HEADER_FILL,
               alignment=HEADER_ALIGNMENT, border=THIN_BORDER)
    for j, label in enumerate(labels):
        cell = ws.cell(row=row, column=2 + j, value=label)
        style_cell(cell, font=HEADER_FONT, fill=HEADER_FILL,
                   alignment=HEADER_ALIGNMENT, border=THIN_BORDER)
    ws.row_dimensions[row].height = 35

    # Data rows
    elements_data = summary.get("elements", {})
    for i, (name, key) in enumerate(zip(ELEMENT_NAMES, ELEMENT_KEYS)):
        row = 4 + i
        cell_a = ws.cell(row=row, column=1, value=name)
        style_cell(cell_a, font=ELEMENT_FONT, fill=ELEMENT_FILL,
                   alignment=Alignment(wrap_text=True, vertical="top"),
                   border=THIN_BORDER)

        values = elements_data.get(key, [])
        max_len = 0
        for j in range(n_papers):
            val = values[j] if j < len(values) else ""
            cell = ws.cell(row=row, column=2 + j, value=val)
            style_cell(cell, font=CONTENT_FONT, alignment=WRAP_ALIGNMENT,
                       border=THIN_BORDER)
            max_len = max(max_len, len(val))

        ws.row_dimensions[row].height = max(45, min(120, max_len * 0.8))


def main():
    parser = argparse.ArgumentParser(description="Export paper analysis to Excel")
    parser.add_argument("--mode", choices=["single", "multi"], required=True)
    parser.add_argument("--output", required=True, help="Output .xlsx path")
    parser.add_argument("--json", required=True, help="JSON data string")
    args = parser.parse_args()

    data = json.loads(args.json)
    wb = Workbook()

    if args.mode == "single":
        ws = wb.active
        ws.title = "论文拆解"
        write_single_sheet(ws, data)
    else:
        papers = data.get("papers", [])
        # Individual sheets
        for i, paper in enumerate(papers):
            if i == 0:
                ws = wb.active
            else:
                ws = wb.create_sheet()
            short_title = paper.get("title", f"论文{i+1}")
            # Sheet name max 31 chars
            ws.title = short_title[:28] + "..." if len(short_title) > 31 else short_title
            write_single_sheet(ws, paper)

        # Summary sheet
        if "summary" in data:
            ws = wb.create_sheet(title="横向对比汇总")
            write_summary_sheet(ws, data["summary"])
            # Move summary sheet to the end (already is)

    wb.save(args.output)
    print(f"Excel saved: {args.output}")


if __name__ == "__main__":
    main()
